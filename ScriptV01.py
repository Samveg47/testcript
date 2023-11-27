from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Load source and target workbooks
source_workbook = load_workbook('Financial_Sample.xlsx')
target_workbook = load_workbook('Book1.xlsx')

# Get specific sheets from the workbooks
source_sheet = source_workbook['Sheet1']  # Change "Sheet1" to your source sheet name
target_sheet = target_workbook['Sheet1']  # Change "Sheet1" to your target sheet name

# Define columns to copy (e.g., columns A, B, C)
columns_to_copy = ['A', 'B', 'C', 'D', 'E', 'F']  # Specify the columns you want to copy

# Find the maximum row with data in the source sheet
max_row = source_sheet.max_row

# Clear existing data in the target sheet
target_sheet.delete_rows(1, target_sheet.max_row)

# Copy data column by column
for col in columns_to_copy:
    column_data = source_sheet[col + '1': col + str(max_row)]  # Get data from the specified column
    for cell in column_data:
        target_sheet[col + str(cell[0].row)] = cell[0].value  # Copy each cell value to the target sheet

# Adjust data in specific cells in the target sheet
target_sheet['A1'] = 'New Value 1'
target_sheet['B2'] = 100
target_sheet['C3'] = '=SUM(A1:B2)'  # You can also use formulas

# Apply text wrapping to all cells in the target sheet with data
for row in target_sheet.iter_rows():
    for cell in row:
        if cell.value:  # Check if cell has data
            cell.alignment = Alignment(wrapText=True)  # Apply text wrapping

# Save the changes to the target workbook
target_workbook.save('Book1.xlsx')
